from __future__ import annotations

import argparse
import re
import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from pypinyin import Style, lazy_pinyin
except ImportError as exc:  # pragma: no cover - depends on local environment
    raise SystemExit(
        "Missing dependency: pypinyin\n"
        "Install it with: python -m pip install pypinyin"
    ) from exc


DEFAULT_ROOT_DIR = Path(r"E:\PCGdata\25_11_2newdata")
DEFAULT_EXCEL = Path(r"F:\PCG data\zhengzhou_PCG.xlsx")
DEFAULT_TARGET_DIR = Path(r"F:\PCG data\dataset\wait2extract")
DEFAULT_SKIP_DIRS = [
    Path(r"F:\PCG data\dataset\test4all_sample"),
    Path(r"F:\PCG data\dataset\no_model"),
    Path(r"F:\PCG data\dataset\no stl"),
]


@dataclass(frozen=True)
class PatientLabel:
    name: str
    row_index: int
    sex: str
    age: str
    primary_disease: str
    symptoms: str
    surgery_date: date
    shunt_type: str
    pvp_pre: str
    pvp_post: str


@dataclass(frozen=True)
class SourceZip:
    path: Path
    ct_date: date
    ct_date_text: str
    chinese_name: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Allocate PCG patient zip files into patient folders with labels."
    )
    parser.add_argument("--root-dir", type=Path, default=DEFAULT_ROOT_DIR)
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--target-dir", type=Path, default=DEFAULT_TARGET_DIR)
    parser.add_argument(
        "--skip-dir",
        type=Path,
        action="append",
        dest="skip_dirs",
        default=None,
        help="Directory whose existing patient folders should be skipped. Can repeat.",
    )
    parser.add_argument(
        "--execute",
        action="store_true",
        default=True,
        help="Actually create folders and copy files. Omit for a dry run.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Process only the first N matching zip files, useful for testing.",
    )
    return parser.parse_args()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def format_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def parse_date_value(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = clean_text(value)
    for fmt in ("%Y%m%d", "%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Cannot parse date: {value!r}")


def name_to_pinyin(name: str) -> str:
    parts = lazy_pinyin(name, style=Style.NORMAL, errors="ignore")
    if not parts:
        raise ValueError(f"Cannot convert name to pinyin: {name!r}")
    return "".join(part.capitalize() for part in parts)


def slug_text(value: str) -> str:
    text = clean_text(value).lower()
    if not text:
        return ""
    pinyin = lazy_pinyin(text, style=Style.NORMAL, errors="ignore")
    slug = "_".join(part for part in pinyin if part)
    return re.sub(r"_+", "_", slug).strip("_")


def translate_sex(value: Any) -> str:
    text = clean_text(value)
    if text in {"1", "1.0", "男", "male"}:
        return "male"
    if text in {"2", "2.0", "女", "female"}:
        return "female"
    return text


def translate_primary_disease(value: Any) -> str:
    text = clean_text(value).replace("，", "")
    if not text:
        return ""

    tokens: list[str] = []
    if "特发性门静脉高压" in text:
        tokens.extend(["idiopathic_portal_hypertension", "portal_hypertension"])
    elif "门静脉高压" in text:
        tokens.append("portal_hypertension")
    if "肝癌" in text:
        tokens.append("hepatocellular_carcinoma")
    if "肝硬化" in text:
        tokens.append("cirrhosis")
    if "食管胃底" in text:
        tokens.append("esophagogastric_variceal_bleeding")
    if "消化道出血" in text or "便血" in text:
        tokens.append("gastrointestinal_bleeding")
    if "门静脉血栓" in text:
        tokens.append("portal_vein_thrombosis")

    return " ".join(dict.fromkeys(tokens)) if tokens else slug_text(text)


def translate_symptoms(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""

    patterns = [
        ("呕血", "hematemesis"),
        ("黑便", "melena"),
        ("便血", "hematochezia"),
        ("血便", "hematochezia"),
        ("腹胀", "abdominal_distension"),
        ("腹痛", "abdominal_pain"),
        ("纳差", "anorexia"),
        ("腹水", "ascites"),
        ("门静脉血栓", "portal_vein_thrombosis"),
    ]
    tokens = [label for needle, label in patterns if needle in text]
    return " ".join(dict.fromkeys(tokens)) if tokens else slug_text(text)


def translate_shunt_type(value: Any) -> str:
    text = clean_text(value).replace(" ", "").replace("，", ",")
    mapping = {
        "门左-肝右": "left_portal_vein_to_right_hepatic_vein",
        "门-肝右静脉": "portal_vein_to_right_hepatic_vein",
        "门-肝右": "portal_vein_to_right_hepatic_vein",
        "门右-肝右": "right_portal_vein_to_right_hepatic_vein",
        "门静脉主干-肝右": "main_portal_vein_to_right_hepatic_vein",
        "门-下腔静脉": "portal_vein_to_inferior_vena_cava",
        "门左-肝中": "left_portal_vein_to_middle_hepatic_vein",
        "门左-下腔静脉": "left_portal_vein_to_inferior_vena_cava",
        "门左-下腔": "left_portal_vein_to_inferior_vena_cava",
        "门-肝左静脉": "portal_vein_to_left_hepatic_vein",
        "肠上-肝右静脉": "superior_mesenteric_vein_to_right_hepatic_vein",
        "肠系膜上静脉-肝右": "superior_mesenteric_vein_to_right_hepatic_vein",
        "门右-下腔静脉": "right_portal_vein_to_inferior_vena_cava",
        "门右-副肝静脉": "right_portal_vein_to_accessory_hepatic_vein",
        "门静脉主干-肝左": "main_portal_vein_to_left_hepatic_vein",
        "门静脉主干-下腔": "main_portal_vein_to_inferior_vena_cava",
        "门静脉主干-下腔静脉": "main_portal_vein_to_inferior_vena_cava",
        "门右-肝中": "right_portal_vein_to_middle_hepatic_vein",
        "teps": "tips",
        "tips": "tips",
        "外院tips后再狭窄": "tips_restenosis",
        "造影": "angiography",
        "": "",
    }
    return mapping.get(text, slug_text(text))


def read_labels(excel_path: Path) -> dict[str, list[PatientLabel]]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    labels: dict[str, list[PatientLabel]] = {}

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name = clean_text(row[0])
        if not name:
            continue
        if row[6] is None:
            print(f"SKIP Excel row {row_index}: missing surgery date for {name}")
            continue

        label = PatientLabel(
            name=name,
            row_index=row_index,
            sex=translate_sex(row[2]),
            age=format_number(row[3]),
            primary_disease=translate_primary_disease(row[4]),
            symptoms=translate_symptoms(row[5]),
            surgery_date=parse_date_value(row[6]),
            shunt_type=translate_shunt_type(row[7]),
            pvp_pre=format_number(row[8]),
            pvp_post=format_number(row[9]),
        )
        labels.setdefault(name, []).append(label)

    return labels


def choose_label(labels: list[PatientLabel], ct_date: date) -> PatientLabel:
    return min(
        labels,
        key=lambda label: (
            abs((ct_date - label.surgery_date).days),
            -label.row_index,
        ),
    )


def iter_source_zips(root_dir: Path) -> list[SourceZip]:
    sources: list[SourceZip] = []
    for path in sorted(root_dir.rglob("*.zip")):
        match = re.match(r"^(\d{8})(.+)$", path.stem)
        if not match:
            print(f"SKIP bad source name: {path}")
            continue
        date_text, chinese_name = match.groups()
        sources.append(
            SourceZip(
                path=path,
                ct_date=parse_date_value(date_text),
                ct_date_text=date_text,
                chinese_name=chinese_name.strip(),
            )
        )
    return sources


def existing_keys(directories: list[Path]) -> set[str]:
    keys: set[str] = set()
    for directory in directories:
        if not directory.exists():
            continue
        for child in directory.iterdir():
            if not child.is_dir():
                continue
            match = re.match(r"^(\d{8})([A-Za-z]+)", child.name)
            if match:
                keys.add((match.group(1) + match.group(2)).lower())
    return keys


def write_text(path: Path, value: str) -> None:
    path.write_text(f"{value}\n", encoding="utf-8")


def create_patient_folder(
    source: SourceZip,
    label: PatientLabel,
    patient_dir: Path,
    post_surgery: bool,
) -> None:
    patient_dir.mkdir(parents=True, exist_ok=False)
    shutil.copy2(source.path, patient_dir / source.path.name)

    label_dir = patient_dir / "label"
    label_dir.mkdir()
    write_text(label_dir / "PVP.txt", label.pvp_post if post_surgery else label.pvp_pre)
    write_text(label_dir / "age.txt", label.age)
    write_text(label_dir / "sex.txt", label.sex)
    write_text(label_dir / "primary_disease.txt", label.primary_disease)
    write_text(label_dir / "symptoms.txt", label.symptoms)
    write_text(label_dir / "surgery_date.txt", label.surgery_date.isoformat())
    write_text(label_dir / "shunt_type.txt", label.shunt_type)


def main() -> int:
    args = parse_args()
    skip_dirs = args.skip_dirs if args.skip_dirs is not None else DEFAULT_SKIP_DIRS

    labels_by_name = read_labels(args.excel)
    sources = iter_source_zips(args.root_dir)
    if args.limit is not None:
        sources = sources[: args.limit]

    skipped_keys = existing_keys([*skip_dirs, args.target_dir])
    created = 0
    skipped_existing = 0
    skipped_no_label = 0
    skipped_no_pvp = 0

    mode = "EXECUTE" if args.execute else "DRY-RUN"
    print(f"Mode: {mode}")
    print(f"Source zip files: {len(sources)}")
    print(f"Known existing patient keys: {len(skipped_keys)}")

    for source in sources:
        labels = labels_by_name.get(source.chinese_name)
        if not labels:
            skipped_no_label += 1
            print(f"SKIP no Excel label: {source.path.name}")
            continue

        label = choose_label(labels, source.ct_date)
        pinyin_name = name_to_pinyin(source.chinese_name)
        key = (source.ct_date_text + pinyin_name).lower()
        if key in skipped_keys:
            skipped_existing += 1
            print(f"SKIP existing: {source.ct_date_text}{pinyin_name} <- {source.path.name}")
            continue

        post_surgery = source.ct_date > label.surgery_date
        patient_name = source.ct_date_text + pinyin_name + ("#" if post_surgery else "")
        patient_dir = args.target_dir / patient_name
        pvp = label.pvp_post if post_surgery else label.pvp_pre
        if not pvp:
            skipped_no_pvp += 1
            phase = "post" if post_surgery else "pre"
            print(
                f"SKIP missing {phase}-surgery PVP: "
                f"{source.path.name}, surgery={label.surgery_date.isoformat()}, row={label.row_index}"
            )
            continue

        print(
            f"CREATE {patient_name}: {source.path.name}, "
            f"surgery={label.surgery_date.isoformat()}, PVP={pvp}, row={label.row_index}"
        )
        if args.execute:
            create_patient_folder(source, label, patient_dir, post_surgery)
            skipped_keys.add(key)
        created += 1

    print(
        f"Summary: create={created}, "
        f"skip_existing={skipped_existing}, "
        f"skip_no_label={skipped_no_label}, skip_no_pvp={skipped_no_pvp}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
